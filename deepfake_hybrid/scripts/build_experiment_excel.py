"""Bangun rekap eksperimen (.xlsx) untuk bukti sidang — detail, dengan kolom
'kenapa train sebelumnya tidak dipakai'. Output ke documents/."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[2] / "sidang" / "Rekap_Eksperimen_Deepfake_2026-07-07.xlsx"

# ── palette ──
NAVY = "2C3540"; INK = "1A1D21"
SPAT = "DDEBF7"; SPAT_T = "1F4E79"
FREQ = "FCE4E1"; FREQ_T = "A02820"
HYB  = "EDE3F5"; HYB_T = "5B3B87"
GREEN = "C6EFCE"; GREEN_T = "1E6B3A"
AMBER = "FFEB9C"; AMBER_T = "8A6100"
GREY = "EEF0F3"; ZEB = "F7F8FA"

thin = Side(style="thin", color="D0D4DA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
H_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
H_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(name="Calibri", bold=True, size=15, color=INK)
SUB_FONT = Font(name="Calibri", italic=True, size=10, color="6A717B")
MONO = Font(name="Consolas", size=10, color=INK)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def dom_fill(model):
    m = (model or "").lower()
    if "spat" in m: return PatternFill("solid", fgColor=SPAT), SPAT_T
    if "freq" in m: return PatternFill("solid", fgColor=FREQ), FREQ_T
    if "hybrid" in m: return PatternFill("solid", fgColor=HYB), HYB_T
    return None, INK


def auc_fill(v):
    try: v = float(str(v).replace(",", "."))
    except Exception: return None
    if v >= 0.85: return PatternFill("solid", fgColor=GREEN)
    if v < 0.62:  return PatternFill("solid", fgColor=AMBER)
    return None


def header(ws, row, cols, widths, title=None, sub=None):
    r = row
    if title:
        ws.cell(r, 1, title).font = TITLE_FONT; r += 1
    if sub:
        ws.cell(r, 1, sub).font = SUB_FONT; r += 1
    r += 1
    for c, (name, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(r, c, name)
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[r].height = 30
    return r + 1


wb = Workbook()

# ══════════════════════════ SHEET 1: JURNAL TRAIN ══════════════════════════
ws = wb.active; ws.title = "1. Jurnal Train"
cols = ["Train", "Tanggal", "Ukuran Sampel", "Perubahan Konfigurasi Utama",
        "Hasil Ringkas (AUC)", "Status", "Alasan tidak dipakai / Catatan"]
w = [7, 15, 16, 40, 34, 15, 46]
hr = header(ws, 1, cols, w,
            "Jurnal Eksperimen — Kronologis (Train 1 → 9)",
            "Bukti seluruh pelatihan yang pernah dijalankan + alasan tiap train sebelumnya tidak dipakai. Sumber: deepfake_hybrid/outputs/*/conclusion.md")
J = [
 ["Train 1","2026-03-12 s/d 03-16","n=50, n=200\n(+ early fusion n=400)",
  "lr 1e-4 · Adam · FreqCNN depth 3 · FFT-norm (x−0,5)/0,5 · 10 epoch · tanpa backbone freeze",
  "Spatial terbaik (FFPP AUC 0,808 @n50). Early fusion (n400) val-AUC puncak 0,685.",
  "TIDAK dipakai",
  "Pipeline & arsitektur belum matang: belum ada AdamW, SE gate, backbone freeze, maupun FreqCNN residual. Hasil belum stabil. Early fusion diuji di sini dan TIDAK lebih baik → tidak dikejar."],
 ["Train 2","2026-03-23 / 24","n=100…1000",
  "FreqCNN 3→5 · + SE gate · + label smoothing 0,05 · + gradient clipping · + noise FFT · scheduler warmup+cosine",
  "Masih fluktuatif; hybrid collapse di beberapa tier.",
  "TIDAK dipakai",
  "Arsitektur baru (dua-cabang + SE gate) tetapi belum stabil; hybrid sering collapse. Masih memakai skema tier lama (300/600/1000)."],
 ["Train 3","2026-03-25","n=100",
  "Adam→AdamW · FreqCNN residual (FreqBlock) · backbone freeze 3 epoch · + ColorJitter/RandErasing · dropout 0,5→0,3",
  "FFPP melompat: S 0,706 · F 0,727 · H 0,787.",
  "TIDAK dipakai",
  "Lompatan besar, tetapi hanya n=100 dan konfigurasi loss belum final (belum ada pos_weight; learning rate belum final)."],
 ["Train 4","2026-04-03","n=100",
  "lr 1e-4→5e-4 · + pos_weight · grad clip 1→5 · spectral mask 0,30→0,15 · label smoothing→0 · tier disatukan 100/250/500/750",
  "CDF sangat baik (S 0,951); FFPP tetap lemah (S 0,542).",
  "TIDAK dipakai",
  "lr 5e-4 terlalu agresif dan label smoothing dimatikan; hanya n=100. Belum konfigurasi akhir."],
 ["Train 5","2026-04-05","n=100, n=250",
  "lr 2e-4 · warmup 2 · patience 10 · seed 0",
  "Regresi di n=250 (turun dari n=100).",
  "TIDAK dipakai",
  "Terjadi regresi di n=250 akibat config-drift (hyperparameter FreqCNN/patience tidak terbawa ke run). Belum stabil antar-tier."],
 ["Train 6","2026-04-09 / 10","n=100",
  "Ablasi: face crop (dengan vs tanpa MTCNN) · separated_ffpp · freq-only (seed 42)",
  "Menetapkan pipeline face crop + isolasi cabang.",
  "Ablasi (bukan hasil utama)",
  "Ini eksperimen ablasi pendukung keputusan pipeline, bukan matriks lengkap untuk hasil utama."],
 ["Train 7","2026-04-15","n=500 (Kaggle)",
  "Full matrix 3 model × 2 dataset",
  "Hybrid-FFPP collapse (early-stop epoch 6); spatial FFPP 0,749; CDF 0,923.",
  "TIDAK dipakai",
  "Hybrid collapse karena early stopping terlalu dini (patience kurang) dan learning rate cabang frekuensi belum seimbang (ikut LR head)."],
 ["Train 8","2026-06-04 (Kaggle)","n=100, n=250",
  "LR rebalance: cabang freq 5e-5 (3-group optimizer) · patience 10→12 · config-drift fix (commit 1985a7a)",
  "Hybrid collapse teratasi; regresi n=250 dibalik (S 0,877).",
  "KONFIGURASI FINAL",
  "Konfigurasi final tervalidasi (collapse & regresi teratasi), tetapi baru sampai n=250. Dipakai sebagai basis konfigurasi; hasil akhir dijalankan di n=750."],
 ["Train 9","FINAL","n=750 · 3 seed",
  "Konfigurasi final + 3 seed (0,1,2) · mean ± std",
  "Spatial terbaik · freq near-random (0,56–0,61) · hybrid tidak > spatial.",
  "✅ DIPAKAI (skripsi)",
  "Satu-satunya run dengan konfigurasi final tervalidasi + ukuran sampel penuh + 3 seed → paling sahih secara ilmiah. Inilah hasil yang dilaporkan di skripsi."],
]
r = hr
for row in J:
    for c, val in enumerate(row, 1):
        cell = ws.cell(r, c, val); cell.border = BORDER; cell.alignment = LEFT
        if c == 1: cell.font = Font(bold=True, color=INK)
        if c == 6:
            if "DIPAKAI" in val and "TIDAK" not in val:
                cell.fill = PatternFill("solid", fgColor=GREEN); cell.font = Font(bold=True, color=GREEN_T)
            elif "FINAL" in val:
                cell.fill = PatternFill("solid", fgColor=SPAT); cell.font = Font(bold=True, color=SPAT_T)
            else:
                cell.fill = PatternFill("solid", fgColor=GREY); cell.font = Font(color="6A717B")
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[r].height = 78
    r += 1
ws.freeze_panes = "A" + str(hr)

# ══════════════════════════ SHEET 2: HASIL DETAIL ══════════════════════════
ws2 = wb.create_sheet("2. Hasil Detail per Train")
cols = ["Train", "Tanggal", "Model", "Dataset / Arah", "Skenario", "AUC", "F1", "Catatan"]
w = [8, 15, 10, 14, 13, 9, 9, 30]
hr2 = header(ws2, 1, cols, w, "Hasil Detail per Train (angka mentah)",
             "Nilai AUC / F1 test-set. Warna model = domain; AUC hijau≥0,85, amber<0,62. Dari conclusion.md tiap folder + tables/n750.")
D = [
 # Train, date, model, ds, scenario, auc, f1, note
 ("Train 1","03-15","spatial","FFPP","in-dataset","0,808","0,751",""),
 ("Train 1","03-15","freq","FFPP","in-dataset","0,541","0,622",""),
 ("Train 1","03-15","hybrid","FFPP","in-dataset","0,662","0,670",""),
 ("Train 1","03-15","spatial","CDF","in-dataset","0,654","0,694",""),
 ("Train 1","03-15","freq","CDF","in-dataset","0,297","0,364","freq sangat lemah"),
 ("Train 1","03-15","hybrid","CDF","in-dataset","0,554","0,602",""),
 ("Train 1","03-16","spatial","FFPP","in-dataset","0,484","0,441","n=200"),
 ("Train 1","03-16","freq","FFPP","in-dataset","0,500","0,440","n=200"),
 ("Train 1","03-16","hybrid","FFPP","in-dataset","0,525","0,536","n=200"),
 ("Train 1","03-16","spatial","CDF","in-dataset","0,710","0,656","n=200"),
 ("Train 1","03-16","freq","CDF","in-dataset","0,691","0,681","n=200"),
 ("Train 1","03-16","hybrid","CDF","in-dataset","0,667","0,585","n=200"),
 ("Train 3","03-25","spatial","FFPP","in-dataset","0,706","0,559","+0,147 vs Train 2"),
 ("Train 3","03-25","freq","FFPP","in-dataset","0,727","0,567",""),
 ("Train 3","03-25","hybrid","FFPP","in-dataset","0,787","0,675","lompatan"),
 ("Train 4","04-03","spatial","CDF","in-dataset","0,951","0,825","CDF sangat baik"),
 ("Train 4","04-03","freq","CDF","in-dataset","0,799","0,676",""),
 ("Train 4","04-03","hybrid","CDF","in-dataset","0,895","0,895",""),
 ("Train 4","04-03","spatial","FFPP","in-dataset","0,542","0,497","FFPP tetap lemah"),
 ("Train 5","04-05","spatial","FFPP","in-dataset","0,552","—","n=250 (regresi)"),
 ("Train 5","04-05","freq","FFPP","in-dataset","0,723","—","n=250"),
 ("Train 5","04-05","hybrid","FFPP","in-dataset","0,563","—","n=250"),
 ("Train 5","04-05","spatial","CDF","in-dataset","0,684","—","n=250"),
 ("Train 5","04-05","freq","CDF","in-dataset","0,578","—","n=250"),
 ("Train 5","04-05","hybrid","CDF","in-dataset","0,575","—","n=250"),
 ("Train 7","04-15","spatial","FFPP","in-dataset","0,749","0,671","n=500"),
 ("Train 7","04-15","freq","FFPP","in-dataset","0,531","0,623","n=500"),
 ("Train 7","04-15","hybrid","FFPP","in-dataset","0,555","0,497","collapse (early-stop ep6)"),
 ("Train 7","04-15","spatial","CDF","in-dataset","0,923","0,861","n=500"),
 ("Train 7","04-15","freq","CDF","in-dataset","0,625","0,595","n=500"),
 ("Train 7","04-15","hybrid","CDF","in-dataset","0,808","0,740","n=500"),
 ("Train 8","06-04","spatial","FFPP","in-dataset","0,877","0,796","n=250 (final config)"),
 ("Train 8","06-04","freq","FFPP","in-dataset","0,670","0,629","n=250"),
 ("Train 8","06-04","hybrid","FFPP","in-dataset","0,668","0,646","collapse teratasi"),
 ("Train 8","06-04","spatial","CDF","in-dataset","0,884","0,794","n=250"),
 ("Train 8","06-04","freq","CDF","in-dataset","0,622","0,651","n=250"),
 ("Train 8","06-04","hybrid","CDF","in-dataset","0,803","0,719","n=250"),
 ("Train 9","FINAL","spatial","FFPP","in-dataset","0,778","0,705","n=750 · 3 seed"),
 ("Train 9","FINAL","freq","FFPP","in-dataset","0,562","0,550","near-random"),
 ("Train 9","FINAL","hybrid","FFPP","in-dataset","0,644","0,606",""),
 ("Train 9","FINAL","spatial","CDF","in-dataset","0,971","0,906","tertinggi"),
 ("Train 9","FINAL","freq","CDF","in-dataset","0,562","0,510","near-random"),
 ("Train 9","FINAL","hybrid","CDF","in-dataset","0,919","0,834",""),
 ("Train 9","FINAL","spatial","FFPP→CDF","cross-dataset","0,678","0,614",""),
 ("Train 9","FINAL","spatial","CDF→FFPP","cross-dataset","0,607","0,137","recall runtuh 0,074"),
 ("Train 9","FINAL","freq","FFPP→CDF","cross-dataset","0,606","0,115","recall 0,064"),
 ("Train 9","FINAL","freq","CDF→FFPP","cross-dataset","0,575","0,526",""),
 ("Train 9","FINAL","hybrid","FFPP→CDF","cross-dataset","0,665","0,594",""),
 ("Train 9","FINAL","hybrid","CDF→FFPP","cross-dataset","0,555","0,238","recall 0,142"),
]
r = hr2
for i, row in enumerate(D):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(r, c, val); cell.border = BORDER
        cell.alignment = CTR if c in (5,6,7) else (LEFT if c in (8,) else Alignment(vertical="center", horizontal="center"))
        if i % 2: cell.fill = PatternFill("solid", fgColor=ZEB)
        if c == 3:
            f, t = dom_fill(val)
            if f: cell.fill = f; cell.font = Font(bold=True, color=t)
        if c == 6:
            cell.font = MONO
            af = auc_fill(val)
            if af: cell.fill = af
    r += 1
ws2.freeze_panes = "A" + str(hr2)

# ══════════════════════════ SHEET 3: HASIL FINAL n750 ══════════════════════════
ws3 = wb.create_sheet("3. Hasil Final (n750)")
cols = ["Model","Latih","Uji","Skenario","Acc","Prec","Recall","F1","AUC","AUC ± std"]
w = [10,8,8,14,8,8,8,8,9,12]
hr3 = header(ws3, 1, cols, w, "Hasil Final — n=750, rata-rata 3 seed (yang dipakai di skripsi)",
             "Sumber: outputs/tables/n750/Table{1,2}_*summary.csv. AUC utama.")
F = [
 ("spatial","FFPP","FFPP","in-dataset","0,700","0,666","0,750","0,705","0,778","± 0,010"),
 ("freq","FFPP","FFPP","in-dataset","0,544","0,525","0,590","0,550","0,562","± 0,007"),
 ("hybrid","FFPP","FFPP","in-dataset","0,600","0,576","0,644","0,606","0,644","± 0,009"),
 ("spatial","CDF","CDF","in-dataset","0,913","0,908","0,904","0,906","0,971","± 0,002"),
 ("freq","CDF","CDF","in-dataset","0,552","0,521","0,503","0,510","0,562","± 0,014"),
 ("hybrid","CDF","CDF","in-dataset","0,847","0,845","0,824","0,834","0,919","± 0,010"),
 ("spatial","FFPP","CDF","cross FFPP→CDF","0,627","0,594","0,637","0,614","0,678","± 0,008"),
 ("freq","FFPP","CDF","cross FFPP→CDF","0,541","0,566","0,064","0,115","0,606","± 0,009"),
 ("hybrid","FFPP","CDF","cross FFPP→CDF","0,624","0,599","0,599","0,594","0,665","± 0,016"),
 ("spatial","CDF","FFPP","cross CDF→FFPP","0,554","0,923","0,074","0,137","0,607","± 0,020"),
 ("freq","CDF","FFPP","cross CDF→FFPP","0,558","0,543","0,531","0,526","0,575","± 0,012"),
 ("hybrid","CDF","FFPP","cross CDF→FFPP","0,564","0,736","0,142","0,238","0,555","± 0,031"),
]
r = hr3
for i, row in enumerate(F):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(r, c, val); cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if i % 2: cell.fill = PatternFill("solid", fgColor=ZEB)
        if c >= 5: cell.font = MONO
        if c == 1:
            f, t = dom_fill(val)
            if f: cell.fill = f; cell.font = Font(bold=True, color=t)
        if c == 9:
            af = auc_fill(val)
            if af: cell.fill = af
    r += 1
ws3.freeze_panes = "A" + str(hr3)

# ══════════════════════════ SHEET 4: EARLY FUSION ══════════════════════════
ws4 = wb.create_sheet("4. Early Fusion (bukti)")
ws4.cell(1,1,"Early Fusion — bukti 'pernah dicoba'").font = TITLE_FONT
ws4.cell(2,1,"Run: outputs/runs/early_fusion_FFPP_n400_seed0/train.log · 2026-03-12 · n=400 · FFPP · seed 0").font = SUB_FONT
ws4.cell(3,1,"XceptionNet 4-kanal (RGB+FFT). Puncak val-AUC ≈ 0,685 → setara/di bawah hybrid → tidak dikejar (jadi 'alternatif konseptual' di BAB II).").font = Font(size=10, color=INK)
ws4.merge_cells("A3:D3"); ws4.cell(3,1).alignment = LEFT; ws4.row_dimensions[3].height = 30
hr4 = header(ws4, 4, ["Epoch","val-AUC","val-F1","Loss"], [10,12,12,12])
EF = [(1,"0,668","0,618","0,139"),(2,"0,639","0,578","0,067"),(3,"0,645","0,587","0,057"),
      (4,"0,657","0,617","0,054"),(5,"0,641","0,591","0,050"),(6,"0,685","0,614","0,043"),
      (7,"0,650","0,624","0,043"),(8,"0,585","0,552","0,029"),(9,"0,628","0,550","0,028")]
r = hr4
for i, row in enumerate(EF):
    for c, val in enumerate(row, 1):
        cell = ws4.cell(r, c, val); cell.border = BORDER; cell.alignment = CTR; cell.font = MONO
        if i % 2: cell.fill = PatternFill("solid", fgColor=ZEB)
        if row[0] == 6: cell.fill = PatternFill("solid", fgColor=GREEN)
    r += 1
ws4.cell(r+1,1,"→ Puncak epoch 6 (val-AUC 0,685). Bandingkan hybrid final (FFPP 0,644 in-dataset) — early fusion tidak memberi keunggulan.").font = Font(italic=True, size=10, color=AMBER_T)

# ══════════════════════════ SHEET 5: FREQBENCH ══════════════════════════
ws5 = wb.create_sheet("5. Freqbench (ResNet18)")
cols = ["Skenario (n=750)","FreqCNN — AUC","FreqCNN — F1","ResNet18 pretrained — AUC","ResNet18 pretrained — F1","ResNet18 scratch — AUC","ResNet18 scratch — F1"]
w = [18,14,13,20,18,18,16]
hr5 = header(ws5, 1, cols, w, "freqbench — backbone alternatif di cabang frekuensi",
             "Bukti: kelemahan frekuensi BUKAN salah FreqCNN. ResNet18 (pretrained/scratch) di input FFT juga near-random. Sumber: roc_cm/freqbench_*_metrics.json")
FB = [
 ("FFPP in-dataset","0,569","0,595","0,561","0,630","0,507","0,511"),
 ("CDF in-dataset","0,578","0,540","0,609","0,469","0,545","0,324"),
 ("FFPP→CDF","0,614","0,101","0,628","0,589","0,544","0,432"),
 ("CDF→FFPP","0,586","0,584","0,578","0,599","0,545","0,345"),
]
r = hr5
for i, row in enumerate(FB):
    for c, val in enumerate(row, 1):
        cell = ws5.cell(r, c, val); cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if i % 2: cell.fill = PatternFill("solid", fgColor=ZEB)
        if c > 1: cell.font = MONO
        if c in (2,4,6):
            af = auc_fill(val)
            if af: cell.fill = af
    r += 1
ws5.cell(r+1,1,"→ Semua backbone frekuensi near-random (AUC 0,51–0,63), termasuk ResNet18-pretrained. Keterbatasan pada representasi FFT, bukan FreqCNN.").font = Font(italic=True, size=10, color=FREQ_T)
ws5.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=7)

# ══════════════════════════ SHEET 6: PARAMETER ══════════════════════════
ws6 = wb.create_sheet("6. Parameter Final")
cols = ["Kategori","Parameter","Nilai"]
w = [22, 30, 40]
hr6 = header(ws6, 1, cols, w, "Parameter & Konfigurasi Final (yang dipakai di Train 9)",
             "Sumber kebenaran: deepfake_hybrid/config.yaml + src/")
P = [
 ("Data","Sampling frame","5 FPS · maks 50 frame/video"),
 ("Data","Ukuran citra","224 × 224"),
 ("Data","Deteksi wajah","MTCNN · margin 0,3 · fallback frame penuh"),
 ("Data","Skip frame hitam","mean piksel < 3"),
 ("Data","Split","70% / 15% / 15% · stratified per-video"),
 ("Data","Rasio kelas","50 : 50 (real : fake)"),
 ("FFT","Grayscale","Y = 0,299R + 0,587G + 0,114B"),
 ("FFT","Transform","fft2 → fftshift → magnitude"),
 ("FFT","High-pass (β)","Gaussian · cutoff 0,15"),
 ("FFT","Log scaling","log(1 + |F(u,v)|)"),
 ("FFT","Normalisasi","z-score per-dataset (fft_stats.json)"),
 ("Augmentasi (RGB)","Resize / crop","256 → RandomResizedCrop 224 (0,8–1,0)"),
 ("Augmentasi (RGB)","ColorJitter","0,2 · 0,2 · 0,1 · 0,05"),
 ("Augmentasi (RGB)","HFlip / RandomErasing","p=0,5 / p=0,1"),
 ("Augmentasi (FFT)","Gaussian noise σ","0,05"),
 ("Augmentasi (FFT)","Spectral band mask","p=0,05"),
 ("Arsitektur — spatial","XceptionNet","timm · ImageNet · fitur 2048-d · ~22 juta param"),
 ("Arsitektur — freq","FreqCNN","depth 5 · base 64 · [64,128,256,512,512] · 512-d · ~4,2 juta"),
 ("Arsitektur — hybrid","Proyeksi / fusi","2048→256 & 512→256 · concat 512 (late fusion)"),
 ("Arsitektur — hybrid","SE gate","512→512 · reduction 4"),
 ("Arsitektur — hybrid","Classifier","Dropout0,5 · 512→128 · ReLU · Dropout0,5 · 128→1"),
 ("Pelatihan","Optimizer","AdamW"),
 ("Pelatihan","Learning rate","base 2e-4 · backbone 2e-5 (÷10) · freq 5e-5 (×0,25) · wd 1e-4"),
 ("Pelatihan","Backbone freeze","3 epoch → unfreeze epoch 4"),
 ("Pelatihan","LR schedule","warmup 3 epoch → cosine (eta_min 1e-6)"),
 ("Pelatihan","Loss","BCEWithLogitsLoss + pos_weight · label smoothing 0,05"),
 ("Pelatihan","Grad clip / accum","max_norm 5,0 / accum 2 (batch efektif 32)"),
 ("Pelatihan","Batch / epoch","16 / maks 30"),
 ("Pelatihan","Early stopping","AUC validasi · patience 12"),
 ("Matriks","Model × Dataset × Sampel × Seed","3 × 2 × 4 × 3 = 72 pelatihan · 144 evaluasi"),
 ("Matriks","Ukuran sampel","100 · 250 · 500 · 750"),
 ("Evaluasi","Metrik","accuracy · precision · recall · F1 · AUC-ROC"),
 ("Evaluasi","Threshold","0,5 + optimal Youden's J"),
 ("Evaluasi","Metrik utama","AUC (seleksi model + early stopping)"),
]
r = hr6
prev = None
for i, row in enumerate(P):
    for c, val in enumerate(row, 1):
        cell = ws6.cell(r, c, val); cell.border = BORDER; cell.alignment = LEFT
        if i % 2: cell.fill = PatternFill("solid", fgColor=ZEB)
        if c == 1:
            cell.font = Font(bold=True, color=INK, size=10)
            f, t = dom_fill(val)
            if f: cell.fill = f; cell.font = Font(bold=True, color=t, size=10)
        if c == 3: cell.font = MONO
    r += 1
ws6.freeze_panes = "A" + str(hr6)

wb.save(OUT)
print("saved ->", OUT)
